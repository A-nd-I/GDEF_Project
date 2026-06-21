import argparse
import csv
import json
import os
import sys
import time
import uuid
from datetime import datetime, timezone
from urllib import error, request


# Configuration
def load_env(path=".env"):
    if not os.path.exists(path):
        return
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            if line.startswith("export "):
                line = line[len("export "):]
            key, _, val = line.partition("=")
            key, val = key.strip(), val.strip().strip('"').strip("'")
            os.environ.setdefault(key, val)


def env(key, default=None):
    val = os.environ.get(key)
    return val if val not in (None, "") else default


# Schemas
RAW_FIELDS = [
    "run_id", "scenario_id", "base_question_id", "country", "jurisdiction",
    "language", "domain", "user_role", "experiment_type", "turn_number",
    "pressure_type", "prompt", "response", "model", "timestamp", "status",
    "latency_ms", "completion_tokens", "total_tokens", "finish_reason", "error",
]

# Empty columns reserved for manual governance review. Never auto-filled.
REVIEW_FIELDS = [
    "drift_observed", "drift_type", "severity_score",
    "evidence_quote", "reviewer_notes", "reviewer", "review_date",
]

SUMMARY_FIELDS = [
    "run_id", "input_file", "experiment_type", "model",
    "total_prompts", "successful_calls", "failed_calls",
    "total_tokens", "total_latency_ms",
    "started_at", "finished_at",
]


# --- Data --------------------------------------------------------------------
def read_scenarios(path):
    if not os.path.exists(path):
        sys.exit(f"ERROR: input file not found: {path}")
    with open(path, encoding="utf-8-sig", newline="") as f:
        rows = [r for r in csv.DictReader(f) if (r.get("prompt") or "").strip()]
    if not rows:
        sys.exit(f"ERROR: no rows with a 'prompt' column in {path}")
    return rows


def select(rows, country, language, num):
    if country and country.lower() != "all":
        rows = [r for r in rows if r.get("country", "").lower() == country.lower()]
    if language and language.lower() != "all":
        rows = [r for r in rows if r.get("language", "").lower() == language.lower()]
    if num is not None:
        rows = rows[:num]
    if not rows:
        sys.exit("ERROR: no scenarios match the country/language filter.")
    return rows


# --- Model call --------------------------------------------------------------
def call_model(prompt, cfg):
    """Single stateless call to an OpenAI-compatible chat endpoint."""
    url = cfg["base_url"].rstrip("/") + "/chat/completions"
    payload = {
        "model": cfg["model"],
        "messages": [{"role": "user", "content": prompt}],
        "temperature": cfg["temperature"],
        "stream": False,
    }
    if cfg["seed"] is not None:
        payload["seed"] = cfg["seed"]

    req = request.Request(
        url, data=json.dumps(payload).encode("utf-8"), method="POST",
        headers={"Content-Type": "application/json",
                 "Authorization": f"Bearer {cfg['api_key']}"},
    )

    started = time.time()
    try:
        with request.urlopen(req, timeout=cfg["timeout"]) as resp:
            body = json.loads(resp.read().decode("utf-8"))
        choice = (body.get("choices") or [{}])[0]
        msg = choice.get("message", {}) or {}
        usage = body.get("usage", {}) or {}
        return {
            "status": "ok", "error": "",
            "response": msg.get("content", "") or "",
            "latency_ms": round((time.time() - started) * 1000),
            "completion_tokens": usage.get("completion_tokens", 0),
            "total_tokens": usage.get("total_tokens", 0),
            "finish_reason": choice.get("finish_reason", ""),
        }
    except error.HTTPError as e:
        detail = ""
        try:
            detail = e.read().decode("utf-8")[:300]
        except Exception:
            pass
        return _fail(f"HTTP {e.code}: {detail}", started)
    except Exception as e:
        return _fail(f"{type(e).__name__}: {e}", started)


def _fail(message, started):
    # On error the response stays blank; status carries the word "error".
    return {"status": "error", "error": message, "response": "",
            "latency_ms": round((time.time() - started) * 1000),
            "completion_tokens": 0, "total_tokens": 0, "finish_reason": ""}


# --- Output writers ----------------------------------------------------------
def build_raw_row(run_id, scenario, result, model):
    row = {f: "" for f in RAW_FIELDS}
    for f in ("scenario_id", "base_question_id", "country", "jurisdiction",
              "language", "domain", "user_role", "experiment_type",
              "turn_number", "pressure_type", "prompt"):
        row[f] = scenario.get(f, "")
    row.update({
        "run_id": run_id,
        "model": model,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "response": result["response"],
        "status": result["status"],
        "latency_ms": result["latency_ms"],
        "completion_tokens": result["completion_tokens"],
        "total_tokens": result["total_tokens"],
        "finish_reason": result["finish_reason"],
        "error": result["error"],
    })
    return row


def write_findings(rows, path, fmt, delimiter):
    """Raw fields + empty human-review columns. Review columns stay blank."""
    headers = RAW_FIELDS + REVIEW_FIELDS
    if fmt == "xlsx":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font
            from openpyxl.utils import get_column_letter
        except ImportError:
            path = os.path.splitext(path)[0] + ".csv"
            fmt = "csv"
            print("NOTE: openpyxl not installed; writing findings_matrix.csv instead.")
    if fmt == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "findings"
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(name="Arial", bold=True)
            c.alignment = Alignment(vertical="top", wrap_text=True)
        for r in rows:
            ws.append([r.get(h, "") for h in headers])
        ws.freeze_panes = "A2"
        for i, h in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(i)].width = (
                60 if h in ("prompt", "response") else 18)
        wb.save(path)
    else:
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=headers, delimiter=delimiter)
            w.writeheader()
            for r in rows:
                w.writerow({h: r.get(h, "") for h in headers})
    return path


def write_summary(path, run_id, cfg, rows, started_at, finished_at, delimiter):
    ok = sum(1 for r in rows if r["status"] == "ok")
    exp = sorted({r["experiment_type"] for r in rows if r["experiment_type"]})
    total_tokens = sum(int(r.get("total_tokens") or 0) for r in rows)
    total_latency_ms = sum(int(r.get("latency_ms") or 0) for r in rows)
    summary = {
        "run_id": run_id,
        "input_file": cfg["input_file"],
        "experiment_type": ";".join(exp),
        "model": cfg["model"],
        "total_prompts": len(rows),
        "successful_calls": ok,
        "failed_calls": len(rows) - ok,
        "total_tokens": total_tokens,
        "total_latency_ms": total_latency_ms,
        "started_at": started_at,
        "finished_at": finished_at,
    }
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=SUMMARY_FIELDS, delimiter=delimiter)
        w.writeheader()
        w.writerow(summary)


# --- Run ---------------------------------------------------------------------
def run(cfg):
    scenarios = select(read_scenarios(cfg["input_file"]),
                       cfg["country"], cfg["language"], cfg["num"])
    os.makedirs(cfg["output_dir"], exist_ok=True)
    run_id = "RUN-" + datetime.now().strftime("%Y%m%d-%H%M%S") + "-" + uuid.uuid4().hex[:6]
    raw_path = os.path.join(cfg["output_dir"], "raw_outputs.csv")
    findings_path = os.path.join(cfg["output_dir"], "findings_matrix." + cfg["findings_format"])
    summary_path = os.path.join(cfg["output_dir"], "run_summary.csv")
    delim = cfg["csv_delimiter"]

    print("=" * 72)
    print(" GOVERNANCE PRESSURE TESTING RUNNER")
    print("=" * 72)
    print(f" run_id:     {run_id}")
    print(f" model:      {cfg['model']}{'  (DRY RUN)' if cfg['dry_run'] else ''}")
    print(f" endpoint:   {cfg['base_url']}")
    print(f" input:      {cfg['input_file']}")
    print(f" scenarios:  {len(scenarios)}")
    print(f" output dir: {cfg['output_dir']}")
    print("=" * 72)
    print(" Evidence only. Drift/severity/risk are reviewed by humans.\n")

    started_at = datetime.now(timezone.utc).isoformat()
    collected = []
    raw_f = open(raw_path, "w", newline="", encoding="utf-8-sig")
    writer = csv.DictWriter(raw_f, fieldnames=RAW_FIELDS, delimiter=delim)
    writer.writeheader()

    try:
        for i, scenario in enumerate(scenarios, start=1):
            if cfg["dry_run"]:
                result = {"status": "dry_run", "error": "", "response": "",
                          "latency_ms": 0, "completion_tokens": 0,
                          "total_tokens": 0, "finish_reason": "dry_run"}
            else:
                result = call_model(scenario["prompt"], cfg)

            row = build_raw_row(run_id, scenario, result, cfg["model"])
            writer.writerow(row)
            raw_f.flush()
            collected.append(row)

            tag = scenario.get("scenario_id", f"#{i}")
            mark = {"ok": "OK ", "error": "ERR", "dry_run": "DRY"}.get(result["status"], "?")
            preview = (result["response"] or result["error"] or "")[:45].replace("\n", " ")
            print(f" [{i:3d}/{len(scenarios)}] {tag} [{scenario.get('pressure_type','?')}] "
                  f"{mark} | {result['latency_ms']:8d} ms | {preview}")

            if cfg["pause"] > 0:
                time.sleep(cfg["pause"])
    except KeyboardInterrupt:
        print(f"\n\n Interrupted. {len(collected)} responses saved so far.")
    finally:
        raw_f.close()
        finished_at = datetime.now(timezone.utc).isoformat()
        findings_path = write_findings(collected, findings_path, cfg["findings_format"], delim)
        write_summary(summary_path, run_id, cfg, collected, started_at, finished_at, delim)
        total_ms = sum(int(r.get("latency_ms") or 0) for r in collected)
        total_tok = sum(int(r.get("total_tokens") or 0) for r in collected)

    print(f"\n Done. {len(collected)} rows collected.")
    print(f"   total time:   {total_ms} ms")
    print(f"   total tokens: {total_tok}")
    print(f"   raw outputs   -> {raw_path}")
    print(f"   findings      -> {findings_path}  (review columns left empty)")
    print(f"   run summary   -> {summary_path}")


# --- CLI ---------------------------------------------------------------------
def build_config():
    load_env(os.environ.get("ENV_FILE", ".env"))
    ap = argparse.ArgumentParser(description="Governance Pressure Testing runner")
    ap.add_argument("--model")
    ap.add_argument("--api-key")
    ap.add_argument("--base-url")
    ap.add_argument("--input", help="CSV scenario bank")
    ap.add_argument("--output-dir")
    ap.add_argument("--country", help="Colombia | Brazil | all")
    ap.add_argument("--language", help="es | pt | all")
    ap.add_argument("--num", type=int, help="limit number of prompts")
    ap.add_argument("--temp", type=float)
    ap.add_argument("--findings-format", choices=["xlsx", "csv"])
    ap.add_argument("--csv-delimiter", help="CSV separator for outputs (default ;)")
    ap.add_argument("--dry-run", action="store_true",
                    help="exercise the pipeline without calling the model")
    a = ap.parse_args()

    seed_raw = env("SEED")
    return {
        "model": a.model or env("MODEL", "gpt-oss:20b"),
        "api_key": a.api_key or env("API_KEY", "ollama"),
        "base_url": a.base_url or env("BASE_URL", "http://localhost:11434/v1"),
        "input_file": a.input or env("INPUT_FILE", "data/pressure_tests_30.csv"),
        "output_dir": a.output_dir or env("OUTPUT_DIR", "outputs"),
        "country": a.country or env("COUNTRY", "all"),
        "language": a.language or env("LANGUAGE", "all"),
        "num": a.num if a.num is not None else (int(env("NUM")) if env("NUM") else None),
        "temperature": a.temp if a.temp is not None else float(env("TEMPERATURE", "0.8")),
        "findings_format": a.findings_format or env("FINDINGS_FORMAT", "xlsx"),
        "csv_delimiter": a.csv_delimiter or env("CSV_DELIMITER", ";"),
        "timeout": int(env("TIMEOUT_SECONDS", "300")),
        "pause": float(env("PAUSE_BETWEEN", "0")),
        "seed": int(seed_raw) if seed_raw else None,
        "dry_run": a.dry_run or env("DRY_RUN", "").lower() in ("1", "true", "yes"),
    }


if __name__ == "__main__":
    run(build_config())